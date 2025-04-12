''' 
This is a file to collect sensor and image data
Jana Armouti (jarmouti)

Based on main.py
'''
import os
import time
import paho.mqtt.client as mqtt
import json
import sys
from datetime import datetime, timedelta
from plant_health.main import health_check
from plant_id_api import identify_plant
from gpiozero import OutputDevice
import serial
import stream
from stream import picam2
from PIL import Image
import io
import RPi.GPIO as GPIO
import dht11
from pymodbus.client import ModbusSerialClient

libdir = os.path.join(os.path.dirname(os.path.dirname(os.path.realpath(__file__))), 'lib')
if os.path.exists(libdir):
  sys.path.append(libdir)

from waveshare_TSL2591 import TSL2591

class Relay(OutputDevice):
  def __init__(self, pin, active_high=False):
    super().__init__(pin, active_high=active_high)

# MQTT configuration
MQTT_SERVER = "broker.emqx.io"  # use your broker address
MQTT_PORT = 1883
MQTT_KEEPALIVE_INTERVAL = 60
MQTT_TOPIC = "django/sproutly/mqtt"  # topic to send data to
CONTROL_TOPIC = "django/sproutly/control" # topic to receive control commands from web app
HEATER_RELAY_PIN = 23
WATER_PUMP_RELAY_PIN = 18
LED_1_RELAY_PIN = 26
LED_2_RELAY_PIN = 6
LED_3_RELAY_PIN = 5
LED_4_RELAY_PIN = 19
WHITE_LIGHT_RELAY_PIN = 16

# start the stream, keep track of if live streaming or not
# stream.start_stream()
streaming = False

HEALTH = "healthy"

# collect sensor data once every 15 minutes
last_sensor_send_time = datetime.now() - timedelta(seconds=885)

# collect image data once every 15 minutes
last_image_send_time = datetime.now() - timedelta(seconds=885)

# collect image data once every 15 minutes
last_led_time = datetime.now() - timedelta(hours=1)

# reset serial buffer data every 2.3 seconds
last_reset_time = datetime.now() - timedelta(seconds=2.3)

# DHT11 sensor
GPIO.setwarnings(False)
GPIO.setmode(GPIO.BCM)
dht_instance = dht11.DHT11(pin=17)

# soil moisture and light sensor data from arduino
ser = serial.Serial('/dev/ttyACM0', 115200, timeout=1)
ser.reset_input_buffer()

# 7-in-1 soil sensor 
soil_client = ModbusSerialClient(
  port="/dev/ttyUSB0",                   # Serial port for rpi
  baudrate=9600,                         # Baudrate for communication
  bytesize=8,                            # Number of bits per byte
  parity="N",                            # No parity
  stopbits=1,                            # One stop bit
)

soil_connected = soil_client.connect()
if soil_connected:
  print("Connected to soil sensor.")
else:
  print("Failed to connect to soil sensor.")

# actuators
actuators_status = {
  "heater": "off",
  "water_pump": "off",
  "nutrients_pump": "off",
  "mister": "off",
  "white_light": "off",
  "LED_light": 0,
  "live_stream": "on",
}

# heater 
heater_relay = Relay(HEATER_RELAY_PIN, active_high=False)
heater_relay.off()

# water pump
water_pump_relay = Relay(WATER_PUMP_RELAY_PIN, active_high=False)
water_pump_relay.off()

# LED light
led_1_relay = Relay(LED_1_RELAY_PIN, active_high=False)
led_2_relay = Relay(LED_2_RELAY_PIN, active_high=False)
led_3_relay = Relay(LED_3_RELAY_PIN, active_high=False)
led_4_relay = Relay(LED_4_RELAY_PIN, active_high=False)
white_light_relay = Relay(WHITE_LIGHT_RELAY_PIN, active_high=False)
led_1_relay.off()
led_2_relay.off()
led_3_relay.off()
led_4_relay.off()
white_light_relay.off()

# keep track of LED light state
last_led_state = 0

# control how may leds currently on
def control_leds(num_leds):
  if num_leds > 0:
    led_1_relay.on()
  else:
    led_1_relay.off()

  if num_leds > 1:
    led_2_relay.on()
  else:
    led_2_relay.off()

  if num_leds > 2:
    led_3_relay.on()
  else:
    led_3_relay.off()

  if num_leds > 3:
    led_4_relay.on()
  else:
    led_4_relay.off()

# callback for when the MQTT client connects to the broker
def on_connect(client, userdata, flags, rc):
  if rc == 0:
    print("Connected successfully to MQTT broker.")
    client.subscribe(CONTROL_TOPIC)
  else:
    print(f"Failed to connect to MQTT broker. Return code {rc}")

# callback for receiving control messages
def on_message(client, userdata, msg):
  global last_led_state
  global streaming
  print(f"msg.topics: {msg.topic}")
  try:
    raw_payload = msg.payload.decode()
    print(f"Received control message: {raw_payload}")
    control_command = json.loads(raw_payload)
    if control_command["command"] == "get_plant_health_check":
      send_plant_health(client)
    elif control_command["command"] == "get_plant_id":
      send_plant_id(client)
    elif control_command["command"] == "get_actuators_status":
      send_actuators_status(client)
    elif "actuator" in control_command:
      if control_command["actuator"] == "heater":
        if control_command["command"] == "on":
          heater_relay.on()
          actuators_status["heater"] = "on"
        elif control_command["command"] == "off":
          heater_relay.off()
          actuators_status["heater"] = "off"
      elif control_command["actuator"] == "water_pump":
        if control_command["command"] == "on":
          water_pump_relay.on()
          actuators_status["water_pump"] = "on"
        elif control_command["command"] == "off":
          water_pump_relay.off()
          actuators_status["water_pump"] = "off"
      elif control_command["actuator"] == "white_light":
        if control_command["command"] == "on":
          white_light_relay.on()
          control_leds(0)
          actuators_status["white_light"] = "on"
          actuators_status["LED_light"] = 0
        elif control_command["command"] == "off":
          white_light_relay.off()
          # go back to prev led state 
          control_leds(last_led_state)
          actuators_status["white_light"] = "off"
          actuators_status["LED_light"] = last_led_state
      elif control_command["actuator"] == "LED_light":
        control_leds(control_command["command"])
        last_led_state = control_command["command"]
        actuators_status["LED_light"] = control_command["command"]
        if control_command["command"] > 0:
          white_light_relay.off()
          actuators_status["white_light"] = "off"
      elif control_command["actuator"] == "live_stream":
        if control_command["command"] == "on":
          if not streaming:
            stream.start_stream()
          streaming = True
          actuators_status["live_stream"] = "on"
        elif control_command["command"] == "off":
          if streaming:
            stream.stop_stream()
          streaming = False
          actuators_status["live_stream"] = "off"

  except json.JSONDecodeError as e:
    print("JSON Decode Error:", e)
    print("Invalid JSON received:", raw_payload)


def send_sensor_data(client, sensor_data):
  global last_sensor_send_time
  try:
    # create a JSON object with the sensor data
    payload = json.dumps(sensor_data)

    # publish the data to the MQTT topic
    client.publish(MQTT_TOPIC, payload)
    print("Published sensor data:", payload)

    last_sensor_send_time = datetime.now()

  except Exception as e:
    print(f"Error in sending sensor data: {e}")


def send_plant_health(client):
  global last_health_check_time
  global streaming
  try:
    # turn white light on and wait for 2 seconds for camera to adjust
    control_leds(0)
    white_light_relay.on()
    time.sleep(2)

    if streaming:
      # get frame from stream
      frame = stream.get_latest_frame()
      image = Image.open(io.BytesIO(frame))
    else:
      # capture image
      picam2.start()
      image = picam2.capture_array('main')
      image = Image.fromarray(image)
      image = image.convert('RGB')
      picam2.stop()

    health_status = health_check(image)

    # turn white light off
    white_light_relay.off()
    assert 0 <= last_led_state <= 4
    control_leds(last_led_state)
    
    payload = json.dumps({
        "type": "plant_health",
        "status": health_status
    })
    
    client.publish(MQTT_TOPIC, payload)
    print("Published plant health status:", payload)

    last_health_check_time = datetime.now()

  except Exception as e:
    print(f"Error in health check: {e}")


def send_plant_id(client):
  global streaming
  try:
    # turn white light on and wait for 2 seconds for camera to adjust
    control_leds(0)
    white_light_relay.on()
    time.sleep(2)
    
    if streaming:
      # get frame from stream
      frame = stream.get_latest_frame()
      image_stream = io.BytesIO(frame)
      picam2.capture_file(image_stream, format="jpeg")
      image_stream.seek(0)
    else:
      # capture image
      picam2.start()
      image_stream = io.BytesIO()
      picam2.capture_file(image_stream, format="jpeg")
      image_stream.seek(0)
      picam2.stop()

    files = [
      ('images', ('image.jpg', image_stream, 'image/jpeg'))
    ]
    
    # turn white light off
    white_light_relay.off()
    assert 0 <= last_led_state <= 4
    control_leds(last_led_state)
    
    best_match, common_names = identify_plant(files)
  
    payload = json.dumps({
      "type": "plant_id",
      "best_match": best_match,
      "common_names": common_names
    })
    
    client.publish(MQTT_TOPIC, payload)
    print("Published plant id:", payload)
  
  except Exception as e:
    print(f"Error in plant identification (api): {e}")


def send_actuators_status(client):
  global actuators_status
  try:
    # create a JSON object with the actuator data
    payload = json.dumps(actuators_status)

    # publish the data to the MQTT topic
    client.publish(MQTT_TOPIC, payload)
    print("Published actuator data:", payload)

  except Exception as e:
    print(f"Error in sending actuator data: {e}")


def get_soil_sensor_data(sensor_data):
  ph_response = soil_client.read_holding_registers(address=0x06, count=1, slave=1)
  if not ph_response.isError():
    sensor_data["ph"] = ph_response.registers[0] / 100.0

  temp_moist_response = soil_client.read_holding_registers(address=0x12, count=2, slave=1)
  if not temp_moist_response.isError():
    sensor_data["soil_moisture_1"] = temp_moist_response.registers[0] / 10.0
    sensor_data["soil_temp"] = temp_moist_response.registers[1] / 10.0

  conductivity_response = soil_client.read_holding_registers(address=0x15, count=1, slave=1)
  if not conductivity_response.isError():
    sensor_data["conductivity"] = conductivity_response.registers[0]

  npk_response = soil_client.read_holding_registers(address=0x1e, count=3, slave=1)
  if not npk_response.isError():
    sensor_data["nitrogen"] = npk_response.registers[0]
    sensor_data["phosphorus"] = npk_response.registers[1]
    sensor_data["potassium"] = npk_response.registers[2]
  
  return (sensor_data)

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
def save_sensor_data_to_excel(sensor_data, excel_path, image_path):
  os.makedirs(os.path.dirname(excel_path), exist_ok=True)
  os.makedirs(image_path, exist_ok=True)
  
  if os.path.exists(excel_path):
    workbook = load_workbook(excel_path)
    sheet = workbook.active
  else:
    workbook = Workbook()
    sheet = workbook.active

    headers = ["Timestamp", "Image Path", "Health", "Temperature (°C)", 
               "Temperature (°F)", "Humidity (%)", "Soil Moisture (%)", 
               "Light (lux)", "Soil pH", "Soil Temp (°C)", "Soil Moisture 7-in-1 (RH%)", 
               "Conductivity (µS/cm)", "Nitrogen (mg/kg)", "Phosphorus (mg/kg)", "Potassium (mg/kg)"]
    sheet.append(headers)
  
  image_files = [f for f in os.listdir(image_path) if f.startswith("image_") and f.endswith(".jpg")]
  image_nums = [int(f.split("_")[1].split(".")[0]) for f in image_files if f.split("_")[1].split(".")[0].isdigit()]
  next_image_num = max(image_nums) + 1 if image_nums else 0

  image_filename = f"image_{next_image_num}.jpg"
  image_filepath = os.path.join(image_path, image_filename)

  row = [
    datetime.now().isoformat(),
    image_filepath,
    HEALTH,
    sensor_data['temperature_c'],
    sensor_data['temperature_f'],
    sensor_data['humidity'],
    sensor_data['soil_moisture'],
    sensor_data['lux'],
    sensor_data['ph'],
    sensor_data['soil_temp'],
    sensor_data['soil_moisture_1'],
    sensor_data['conductivity'],
    sensor_data['nitrogen'],
    sensor_data['phosphorus'],
    sensor_data['potassium'],   
  ]
  sheet.append(row)

  workbook.save(excel_path)

  return image_path

def save_image_data(image_path):
  global streaming

  # turn white light on and wait for 2 seconds for camera to adjust
  control_leds(0)
  white_light_relay.on()
  time.sleep(2)

  if streaming:
    # get frame from stream
    frame = stream.get_latest_frame()
    image = Image.open(io.BytesIO(frame))
  else:
    # capture image
    picam2.start()
    image = picam2.capture_array('main')
    image = Image.fromarray(image)
    image = image.convert('RGB')
    picam2.stop()

  image.save(image_path)

  # turn white light off
  white_light_relay.off()
  assert 0 <= last_led_state <= 4
  control_leds(last_led_state)

  return 

    
# initialize MQTT client
client = mqtt.Client()
client.on_connect = on_connect
client.on_message = on_message

client.connect(MQTT_SERVER, MQTT_PORT, MQTT_KEEPALIVE_INTERVAL)
client.loop_start()

sensor_data = {
  "temperature_c": 0,
  "temperature_f": 0,
  "humidity": 0,
  "soil_moisture": 0,
  "lux": 0,
  "ph": 0,
  "soil_moisture_1": 0,
  "soil_temp": 0,
  "conductivity": 0,
  "nitrogen": 0,
  "phosphorus": 0,
  "potassium": 0,
}

try:
  while True:
    try:
      # get dht11 sensor data
      try:
        dht_result = dht_instance.read()
        if dht_result.is_valid():
          sensor_data['temperature_c'] = dht_result.temperature
          sensor_data['temperature_f'] = sensor_data['temperature_c'] * (9 / 5) + 32
          sensor_data['humidity'] = dht_result.humidity
      except RuntimeError as err:
        print(err)

      # get arduino sensor data
      if ser.in_waiting > 0:
        try:
          line = ser.readline().decode('utf-8').strip()
          values = line.split(',')
          if len(values) == 2:
            sensor_data['soil_moisture'] = float(values[0])
            sensor_data['lux'] = int(values[1])
        except ValueError: 
          print(f"Invalid data received")

      # get 7-in-1 soil sensor data
      sensor_data = get_soil_sensor_data(sensor_data)

      print(
      f"Temp: {sensor_data['temperature_c']:.1f}°C / {sensor_data['temperature_f']:.1f}°F | "
      f"Humidity: {sensor_data['humidity']}% | "
      f"Soil Moisture: {sensor_data['soil_moisture']}% | "
      f"Light: {sensor_data['lux']} lux | "
      f"Soil pH: {sensor_data['ph']} pH | "
      f"Soil Temp: {sensor_data['soil_temp']}°C | "
      f"Soil Moisture 7-in-1: {sensor_data['soil_moisture_1']} RH% | "
      f"Conductivity: {sensor_data['conductivity']} µS/cm | "
      f"N: {sensor_data['nitrogen']} mg/kg | "
      f"P: {sensor_data['phosphorus']} mg/kg | "
      f"K: {sensor_data['potassium']} mg/kg"
      )

      # check if 15 minute has passed since last sensor data was collected
      if datetime.now() - last_sensor_send_time >= timedelta(minutes=15):
        # save sensor data to excel file
        excel_path = "/home/sproutly/Desktop/Sproutly/rpi/plant_health/datasets/rpi/sensor_log.xlsx"
        image_path = "/home/sproutly/Desktop/Sproutly/rpi/plant_health/datasets/rpi/images"
        image_filepath = save_sensor_data_to_excel(sensor_data, excel_path, image_path)
        last_sensor_send_time = datetime.now()
        print(f"Updated log: {excel_path}")
      
      # check if 15 minute has passed since last image data was collected
      if datetime.now() - last_image_send_time >= timedelta(minutes=15):
        # save image data to image folder
        save_image_data(image_filepath)
        last_image_send_time = datetime.now()
        print(f"Added image: {image_filepath}")
      
      # check if 2.3 seconds have passed since serial buffer reset
      if datetime.now() - last_reset_time >= timedelta(seconds=2.3):
        ser.reset_input_buffer()
        last_reset_time = datetime.now()

      # for variety in sensor data, change LED state every hour
      if datetime.now() - last_led_time >= timedelta(hours=1):
        last_led_state = (last_led_state + 1) % 4
        control_leds(last_led_state)
        last_led_time = datetime.now()

    except RuntimeError as err:
      print(err.args[0])

    time.sleep(2.0)

except KeyboardInterrupt:
  if soil_connected:
    soil_client.close()
  GPIO.cleanup()